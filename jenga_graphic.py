from openpyxl import load_workbook
import math
import svgwrite
from svgwrite.path import Path as SVGPath
from svgwrite.container import Group
from svgwrite.shapes import Rect
from PIL import ImageFont
import xml.etree.ElementTree as ET
from pathlib import Path
import re

# ----------------------
# Parameters
dir = Path("C:\\Users\\lohoff\\Documents\\Private\\Makerspace\\Jenga\\")
in_file = dir / "Spicy_Jenga_Vorschläge.xlsx"
out_file = dir / "grid.svg"

stone_width = 7.5
stone_height = 2.5

text_height = 1.8
text_width = 5
font_size = 0.7

icon_size = 0.8   # cm
icon_gap = .5    # cm

width_total = 50

green = "FFB6D7A8"
red = "FFEA9999"
icons = {
    red: dir / "chili.svg",
    green: dir / "cloud.svg",
}

# -----------------------
def measure_text_width(text, font_size_cm):
    """
    Measure text width in cm using PIL and Arial.
    """
    CM_TO_PX = 96 / 2.54
    font_size_px = font_size_cm * CM_TO_PX

    font = ImageFont.truetype("arial.ttf", int(font_size_px))
    bbox = font.getbbox(text)  # returns (x0, y0, x1, y1)
    width_px = bbox[2] - bbox[0]
    return width_px / CM_TO_PX  # back to cm

assert stone_width > text_width + icon_size + icon_gap, "Text plus icon would be too wide for the block size"

# Load the sheet
wb = load_workbook(in_file, data_only=True)
ws = wb.active  # use first sheet, or wb["SheetName"]

# Get row texts and colors
snippets = []
for row in ws.iter_rows(values_only=False):  # values_only=False so we also get style
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            # Get the text
            text = cell.value.strip()

            # Get the fill color (RGB hex like "FF00FF00" for green)
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
            snippets.append({
                "text": text,
                "color": fill
            })

# Grid calculations
N = int(width_total // stone_width)
M = math.ceil(len(snippets) / N)
grid_width = N * stone_width
grid_height = M * stone_height

dwg = svgwrite.Drawing(out_file, size=(f"{grid_width}cm", f"{grid_height}cm"))

# Draw grid lines
for i in range(N + 1):
    x = i * stone_width
    dwg.add(dwg.line((f"{x}cm", "0cm"), (f"{x}cm", f"{grid_height}cm"),
                     stroke="green", stroke_width=0.1))

for j in range(M + 1):
    y = j * stone_height
    dwg.add(dwg.line(("0cm", f"{y}cm"), (f"{grid_width}cm", f"{y}cm"),
                     stroke="green", stroke_width=0.1))


def wrap_text_to_box(text, font_size, box_width, box_height):
    """
    Wrap text word by word into lines that fit within box_width.
    If text doesn't fit into box_height with current font_size, return None.
    """
    words = text.split()
    lines, current = [], ""

    for word in words:
        test_line = (current + " " + word).strip()
        est_width = measure_text_width(test_line, font_size)
        if est_width <= box_width:
            current = test_line
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)

    # NEW: check width overflow (important for long single words)
    if any(measure_text_width(line, font_size) > box_width for line in lines):
        return None

    total_height = len(lines) * font_size
    if total_height > box_height:
        return None  # doesn’t fit
    return lines


def load_icon_as_path(filename, max_size):
    """
    Load first <path> from SVG, normalize to (0,0), scale to fit max_size.
    Returns an svgwrite Group ready to insert at (0,0).
    """
    tree = ET.parse(filename)
    root = tree.getroot()

    ns = {"svg": "http://www.w3.org/2000/svg"}
    path_elem = root.find(".//{http://www.w3.org/2000/svg}path")
    if path_elem is None:
        raise ValueError(f"No <path> found in {filename}")

    d = path_elem.attrib["d"]
    path = SVGPath(d=d, fill="black")

    # get viewBox
    viewBox = root.attrib.get("viewBox")
    if viewBox:
        x0, y0, w, h = map(float, viewBox.split())
    else:
        x0, y0, w, h = 0, 0, 100, 100  # fallback

    scale = min(max_size / w, max_size / h)

    g = Group()
    # scale first, then translate to normalize origin
    g.add(path)
    g.scale(scale)
    g.translate(-x0, -y0)

    # DEBUG: bounding box
    g.add(Rect(insert=(0, 0), size=(w*scale, h*scale),
               fill="none", stroke="red", stroke_width=0.05))

    return g, w*scale, h*scale  # also return scaled width/height


def _float_attr(elem, name, default=0.0):
    v = elem.attrib.get(name)
    return float(v) if v is not None else default

def _add_pt_to_bbox(bbox, x, y):
    if bbox is None:
        return [x, y, x, y]
    xmin, ymin, xmax, ymax = bbox
    xmin = min(xmin, x); ymin = min(ymin, y)
    xmax = max(xmax, x); ymax = max(ymax, y)
    return [xmin, ymin, xmax, ymax]

def _merge_bbox(bbox, other):
    if other is None:
        return bbox
    if bbox is None:
        return other[:]
    return [
        min(bbox[0], other[0]),
        min(bbox[1], other[1]),
        max(bbox[2], other[2]),
        max(bbox[3], other[3]),
    ]

# --- Path 'd' parser that produces a bbox (tries to be robust) ---
def path_bbox_from_d(d):
    # tokens: commands and numbers
    tokens = re.findall(r'[A-Za-z]|[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', d)
    if not tokens:
        return None
    idx = 0
    cur_x = 0.0
    cur_y = 0.0
    start_x = 0.0
    start_y = 0.0
    prev_cmd = None
    prev_ctrl_x = None
    prev_ctrl_y = None

    inf = float('inf')
    xmin, ymin, xmax, ymax = inf, inf, -inf, -inf

    def add_pt(x, y):
        nonlocal xmin, ymin, xmax, ymax
        xmin = min(xmin, x); ymin = min(ymin, y)
        xmax = max(xmax, x); ymax = max(ymax, y)

    counts = {'M':2,'L':2,'H':1,'V':1,'C':6,'S':4,'Q':4,'T':2,'A':7,'Z':0}

    while idx < len(tokens):
        token = tokens[idx]
        if re.match(r'[A-Za-z]', token):
            cmd = token
            idx += 1
        else:
            # implicit repeat of previous command
            if prev_cmd is None:
                # invalid path data
                break
            cmd = prev_cmd

        cmd_upper = cmd.upper()
        is_rel = cmd.islower()

        if cmd_upper == 'Z':
            cur_x, cur_y = start_x, start_y
            add_pt(cur_x, cur_y)
            prev_cmd = cmd
            prev_ctrl_x = prev_ctrl_y = None
            continue

        need = counts.get(cmd_upper)
        if need is None:
            # unknown command -> break
            break

        # special treatment for 'M': first pair is moveto, subsequent pairs are lineto
        if cmd_upper == 'M':
            if idx + 1 >= len(tokens):
                break
            x = float(tokens[idx]); y = float(tokens[idx+1]); idx += 2
            if is_rel:
                x += cur_x; y += cur_y
            cur_x, cur_y = x, y
            start_x, start_y = cur_x, cur_y
            add_pt(cur_x, cur_y)
            prev_cmd = cmd
            prev_ctrl_x = prev_ctrl_y = None
            # process subsequent (x,y) pairs as implicit lineto
            while idx + 1 < len(tokens) and not re.match(r'[A-Za-z]', tokens[idx]):
                x = float(tokens[idx]); y = float(tokens[idx+1]); idx += 2
                if is_rel:
                    x += cur_x; y += cur_y
                cur_x, cur_y = x, y
                add_pt(cur_x, cur_y)
                prev_cmd = 'L' if cmd.isupper() else 'l'
                prev_ctrl_x = prev_ctrl_y = None
            continue

        # process repeated segments for the same command
        while idx + need - 1 < len(tokens) and not re.match(r'[A-Za-z]', tokens[idx]):
            if cmd_upper == 'L':
                x = float(tokens[idx]); y = float(tokens[idx+1]); idx += 2
                if is_rel:
                    x += cur_x; y += cur_y
                cur_x, cur_y = x, y
                add_pt(x, y)
                prev_ctrl_x = prev_ctrl_y = None

            elif cmd_upper == 'H':
                x = float(tokens[idx]); idx += 1
                if is_rel:
                    x += cur_x
                cur_x = x
                add_pt(cur_x, cur_y)
                prev_ctrl_x = prev_ctrl_y = None

            elif cmd_upper == 'V':
                y = float(tokens[idx]); idx += 1
                if is_rel:
                    y += cur_y
                cur_y = y
                add_pt(cur_x, cur_y)
                prev_ctrl_x = prev_ctrl_y = None

            elif cmd_upper == 'C':
                x1 = float(tokens[idx]); y1 = float(tokens[idx+1])
                x2 = float(tokens[idx+2]); y2 = float(tokens[idx+3])
                x  = float(tokens[idx+4]); y  = float(tokens[idx+5])
                idx += 6
                if is_rel:
                    x1 += cur_x; y1 += cur_y
                    x2 += cur_x; y2 += cur_y
                    x  += cur_x; y  += cur_y
                add_pt(x1, y1); add_pt(x2, y2); add_pt(x, y)
                prev_ctrl_x, prev_ctrl_y = x2, y2
                cur_x, cur_y = x, y

            elif cmd_upper == 'S':
                x2 = float(tokens[idx]); y2 = float(tokens[idx+1])
                x  = float(tokens[idx+2]); y  = float(tokens[idx+3])
                idx += 4
                if is_rel:
                    x2 += cur_x; y2 += cur_y
                    x  += cur_x; y  += cur_y
                # implied first control = reflection of previous control (if previous was C or S)
                if prev_cmd and prev_cmd.upper() in ('C','S') and prev_ctrl_x is not None:
                    x1 = 2*cur_x - prev_ctrl_x
                    y1 = 2*cur_y - prev_ctrl_y
                    add_pt(x1, y1)
                add_pt(x2, y2); add_pt(x, y)
                prev_ctrl_x, prev_ctrl_y = x2, y2
                cur_x, cur_y = x, y

            elif cmd_upper == 'Q':
                x1 = float(tokens[idx]); y1 = float(tokens[idx+1])
                x  = float(tokens[idx+2]); y  = float(tokens[idx+3])
                idx += 4
                if is_rel:
                    x1 += cur_x; y1 += cur_y
                    x  += cur_x; y  += cur_y
                add_pt(x1, y1); add_pt(x, y)
                prev_ctrl_x, prev_ctrl_y = x1, y1
                cur_x, cur_y = x, y

            elif cmd_upper == 'T':
                x = float(tokens[idx]); y = float(tokens[idx+1]); idx += 2
                if is_rel:
                    x += cur_x; y += cur_y
                # implied control for smooth quadratic
                if prev_cmd and prev_cmd.upper() in ('Q','T') and prev_ctrl_x is not None:
                    x1 = 2*cur_x - prev_ctrl_x
                    y1 = 2*cur_y - prev_ctrl_y
                    add_pt(x1, y1)
                    prev_ctrl_x, prev_ctrl_y = x1, y1
                add_pt(x, y)
                cur_x, cur_y = x, y

            elif cmd_upper == 'A':
                rx = float(tokens[idx]); ry = float(tokens[idx+1])
                rot = float(tokens[idx+2]); large = float(tokens[idx+3])
                sweep = float(tokens[idx+4])
                x = float(tokens[idx+5]); y = float(tokens[idx+6])
                idx += 7
                if is_rel:
                    x += cur_x; y += cur_y
                # approximate: include endpoint and ±rx/ry around endpoint (good approximation for many icons)
                add_pt(x, y)
                add_pt(x + rx, y + ry)
                add_pt(x - rx, y - ry)
                prev_ctrl_x = prev_ctrl_y = None
                cur_x, cur_y = x, y

            else:
                # unknown command, bail
                break

            prev_cmd = cmd

    if xmin == inf:
        return None
    return (xmin, ymin, xmax, ymax)

# --- helpers for other SVG elements ---
def rect_bbox(elem):
    x = _float_attr(elem, 'x', 0.0)
    y = _float_attr(elem, 'y', 0.0)
    w = _float_attr(elem, 'width', 0.0)
    h = _float_attr(elem, 'height', 0.0)
    return (x, y, x + w, y + h)

def circle_bbox(elem):
    cx = _float_attr(elem, 'cx', 0.0)
    cy = _float_attr(elem, 'cy', 0.0)
    r = _float_attr(elem, 'r', 0.0)
    return (cx - r, cy - r, cx + r, cy + r)

def ellipse_bbox(elem):
    cx = _float_attr(elem, 'cx', 0.0)
    cy = _float_attr(elem, 'cy', 0.0)
    rx = _float_attr(elem, 'rx', 0.0)
    ry = _float_attr(elem, 'ry', 0.0)
    return (cx - rx, cy - ry, cx + rx, cy + ry)

def line_bbox(elem):
    x1 = _float_attr(elem, 'x1', 0.0)
    y1 = _float_attr(elem, 'y1', 0.0)
    x2 = _float_attr(elem, 'x2', 0.0)
    y2 = _float_attr(elem, 'y2', 0.0)
    return (min(x1,x2), min(y1,y2), max(x1,x2), max(y1,y2))

def poly_points_bbox(elem):
    pts = elem.attrib.get('points','').strip()
    if not pts:
        return None
    nums = re.findall(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', pts)
    if not nums:
        return None
    xs = list(map(float, nums[0::2]))
    ys = list(map(float, nums[1::2]))
    return (min(xs), min(ys), max(xs), max(ys))

# --- final function that creates the transform based on true icon bounds ---
def create_icon_path(filename, max_size_cm, insert_x_cm, insert_y_cm):
    """Load first path from SVG, compute true content bbox, scale so the LARGEST dimension
       matches max_size_cm, and position top-left at insert_x_cm,insert_y_cm."""
    CM_TO_PX = 96.0 / 2.54

    tree = ET.parse(filename)
    root = tree.getroot()
    ns = {"svg": "http://www.w3.org/2000/svg"}

    # compute bbox from content (paths and basic shapes)
    total_bbox = None

    # paths
    for path_elem in root.findall(".//svg:path", ns):
        d = path_elem.attrib.get('d','')
        if not d:
            continue
        bbox = path_bbox_from_d(d)
        if bbox:
            total_bbox = _merge_bbox(total_bbox, bbox)

    # rect/circle/ellipse/line/polyline/polygon
    for rect in root.findall(".//svg:rect", ns):
        total_bbox = _merge_bbox(total_bbox, rect_bbox(rect))
    for circ in root.findall(".//svg:circle", ns):
        total_bbox = _merge_bbox(total_bbox, circle_bbox(circ))
    for ell in root.findall(".//svg:ellipse", ns):
        total_bbox = _merge_bbox(total_bbox, ellipse_bbox(ell))
    for ln in root.findall(".//svg:line", ns):
        total_bbox = _merge_bbox(total_bbox, line_bbox(ln))
    for poly in root.findall(".//svg:polyline", ns) + root.findall(".//svg:polygon", ns):
        bbox = poly_points_bbox(poly)
        if bbox:
            total_bbox = _merge_bbox(total_bbox, bbox)

    # fallback to viewBox if nothing found
    viewBox = root.attrib.get("viewBox")
    if total_bbox is None:
        if viewBox:
            x0, y0, w, h = map(float, viewBox.split())
            total_bbox = (x0, y0, x0 + w, y0 + h)
        else:
            # last resort fallback
            total_bbox = (0.0, 0.0, 100.0, 100.0)

    xmin, ymin, xmax, ymax = total_bbox
    icon_w = xmax - xmin
    icon_h = ymax - ymin
    if icon_w <= 0 or icon_h <= 0:
        raise ValueError("Computed empty icon bbox")

    max_size_px = max_size_cm * CM_TO_PX
    # choose scale so the LARGER dimension becomes max_size_px
    scale = max_size_px / max(icon_w, icon_h)

    tx_px = insert_x_cm * CM_TO_PX - xmin * scale
    ty_px = insert_y_cm * CM_TO_PX - ymin * scale

    transform_str = f"translate({tx_px:.4f},{ty_px:.4f}) scale({scale:.8f})"

    # find the first path to return (like your original function)
    first_path = root.find(".//svg:path", ns)
    if first_path is None:
        raise ValueError(f"No path found in {filename}")
    d = first_path.attrib.get('d','')
    # you returned SVGPath(...) before — mimic that return
    return SVGPath(d=d, fill="black", transform=transform_str)


# --- Place text snippets ---
for idx, item in enumerate(snippets):
    text = item["text"]
    color = item["color"]

    row = idx // N
    col = idx % N

    # Stone top-left corner
    x0 = col * stone_width
    y0 = row * stone_height
    cx = x0 + stone_width / 2
    cy = y0 + stone_height / 2

    # Start with default size
    this_font_size = font_size
    lines = wrap_text_to_box(text, this_font_size, text_width, text_height)

    # Shrink font until text fits
    while lines is None and this_font_size > 0.2:
        this_font_size *= 0.9
        lines = wrap_text_to_box(text, this_font_size, text_width, text_height)

    # Save chosen font size
    item["font_size"] = this_font_size

    if not lines:
        lines = [text]  # fallback, single line

    # --- Measure text block (unchanged) ---
    text_block_width = max(measure_text_width(line, this_font_size) for line in lines)

    text_block_height = len(lines) * this_font_size

    # --- Compute combined width including icon (if present) ---
    icon_path = None
    if color and color in icons:
        icon_path = icons[color]

    # --- ICON vector handling ---
    icon_group = None
    icon_w, icon_h = 0, 0
    if color and color in icons:
        icon_group, icon_w, icon_h = load_icon_as_path(icons[color], icon_size)

    # Total width including icon
    if icon_group:
        total_width = text_block_width + icon_gap + icon_w
    else:
        total_width = text_block_width

    start_x = cx - total_width/2

    # --- Draw text ---
    start_y = cy - text_block_height/2 + this_font_size/2
    for i, line in enumerate(lines):
        y = start_y + i*this_font_size
        dwg.add(dwg.text(
            line,
            insert=(f"{start_x + text_block_width/2}cm", f"{y}cm"),
            text_anchor="middle",
            dominant_baseline="middle",
            font_size=f"{this_font_size}cm",
            font_family="Arial",
            fill="red"
        ))

    # --- Place icon ---
    if color and color in icons:
        icon_insert_x = start_x + text_block_width + icon_gap
        icon_insert_y = cy - icon_size/2
        icon_path = create_icon_path(icons[color], icon_size, icon_insert_x, icon_insert_y)
        dwg.add(icon_path)

dwg.save(out_file)


